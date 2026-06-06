import openpyxl

wb = openpyxl.load_workbook("Plan_PPL_Dziennik_NEW.xlsx")
print(wb.sheetnames)

ws = wb["PUSH – Klatka · Barki · Triceps"]

for row in ws.iter_rows(8,12):
    print(row[1].value)